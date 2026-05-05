"""
Top 10 Suspect Accounts Module.

Generates a report of top 10 accounts by disputed amount with automatic filtering.
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.column_detector import ColumnDetector
from src.data_processor import DataProcessor
from src.aggregation_engine import AggregationEngine
from src.models import ColumnMapping


def render_top_10_suspect_accounts_page():
    """Render the Top 10 Suspect Accounts page."""
    st.title("🎯 Top 10 Suspect Accounts from Layer 1")
    st.markdown("Automatically generates top 10 accounts by disputed amount with bank filtering")
    
    st.markdown("---")
    
    # File upload
    st.subheader("📤 Upload Layer 1 File")
    uploaded_file = st.file_uploader(
        "Upload Excel file (Layer 1 data)",
        type=['xlsx', 'xls'],
        key="top10_upload",
        help="Upload the Layer 1 file containing account and transaction data"
    )
    
    if not uploaded_file:
        st.info("👆 Please upload a Layer 1 Excel file to generate the report")
        return
    
    # Process immediately after upload
    try:
        with st.spinner("📊 Processing file and generating report..."):
            # Read the file
            df = pd.read_excel(uploaded_file)
            st.success(f"✅ File loaded: {len(df)} rows")
            
            # Initialize services
            column_detector = ColumnDetector()
            data_processor = DataProcessor()
            aggregation_engine = AggregationEngine()
            
            # Auto-detect columns
            st.info("🔍 Auto-detecting columns...")
            mapping = column_detector.detect_columns(df)
            
            if not mapping:
                st.error("❌ Could not auto-detect required columns. Please check your file format.")
                return
            
            st.success(f"✅ Columns detected: Account No., Bank, ACK No., Disputed Amount")
            
            # STEP 1: Filter data
            st.info("🔧 Applying filters...")
            
            # Find column names
            ack_col = None
            bank_col = None
            
            for col in df.columns:
                if 'acknowledgement' in col.lower() or 'ack' in col.lower():
                    ack_col = col
                if 'bank' in col.lower() or 'fi' in col.lower():
                    bank_col = col
            
            if not ack_col or not bank_col:
                st.error("❌ Required columns not found (ACK Number, Bank)")
                return
            
            initial_count = len(df)
            
            # Filter 1: ACK must start with "311"
            df_filtered = df[df[ack_col].astype(str).str.startswith('311', na=False)].copy()
            ack_filtered = initial_count - len(df_filtered)
            
            # Filter 2: Bank name must contain "bank", "Bank", or "BANK"
            df_filtered = df_filtered[
                df_filtered[bank_col].astype(str).str.contains('Bank|bank|BANK', na=False, regex=True)
            ].copy()
            bank_filtered = initial_count - ack_filtered - len(df_filtered)
            
            st.success(f"✅ Filters applied: Removed {ack_filtered} non-311 ACKs, {bank_filtered} non-bank entries")
            st.info(f"📊 Processing {len(df_filtered)} valid records")
            
            # STEP 2: Clean and aggregate
            st.info("🔨 Aggregating by account...")
            cleaned_df = data_processor.clean_dataframe(df_filtered, mapping)
            aggregated = aggregation_engine.aggregate_by_account(cleaned_df, mapping)
            sorted_accounts = aggregation_engine.sort_results(aggregated)
            
            # STEP 3: Get top 10 by disputed amount
            st.info("🎯 Selecting top 10 accounts by disputed amount...")
            
            # Sort by disputed amount (descending) and take top 10
            top_10 = sorted(sorted_accounts, key=lambda x: x.disputed_amount, reverse=True)[:10]
            
            st.success(f"✅ Top 10 accounts selected (Total disputed: ₹{sum(acc.disputed_amount for acc in top_10):,.2f})")
            
            # STEP 4: Generate Excel file
            st.info("📄 Generating Excel report...")
            
            # Get yesterday's date
            yesterday = datetime.now() - timedelta(days=1)
            date_str = yesterday.strftime("%d-%m-%Y")
            filename = f"{date_str} Top 10 Suspect Accounts  from Layer 1.xlsx"
            
            # Create Excel file
            excel_bytes = generate_top_10_excel(top_10, date_str)
            
            st.success("✅ Report generated successfully!")
            
            # Show summary
            st.markdown("---")
            st.subheader("📊 Top 10 Accounts Summary")
            
            # Display summary table
            summary_data = []
            for i, acc in enumerate(top_10, 1):
                summary_data.append({
                    "Rank": i,
                    "Account No.": acc.account_number,
                    "Bank": acc.bank_name,
                    "Disputed Amount": f"₹{acc.disputed_amount:,.2f}",
                    "Total Amount": f"₹{acc.total_amount:,.2f}",
                    "Transactions": acc.transaction_count
                })
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
            
            # Download button
            st.markdown("---")
            st.download_button(
                label="⬇️ Download Excel Report",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
            
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


def generate_top_10_excel(accounts, date_str):
    """
    Generate Excel file for top 10 suspect accounts.
    
    Uses the same format as aggregated by account output.
    
    Args:
        accounts: List of top 10 AggregatedAccount objects
        date_str: Date string for the title
        
    Returns:
        Excel file as bytes
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Top 10 Suspect Accounts"
    
    # Define styles
    title_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    header_font = Font(name='Calibri', size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    normal_font = Font(name='Calibri', size=11, color="000000")
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"{date_str} Top 10 Suspect Accounts from Layer 1"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_align
    title_cell.border = border
    ws.row_dimensions[1].height = 30
    
    # Headers (same as aggregated by account)
    headers = [
        "SR NO",
        "Account No.",
        "Bank Name",
        "IFSC Code",
        "Address",
        "Mobile",
        "Email",
        "Transaction Count",
        "ACK Numbers",
        "Total Amount",
        "Disputed Amount"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    ws.row_dimensions[2].height = 25
    
    # Data rows
    for row_idx, account in enumerate(accounts, 3):
        # SR NO
        cell = ws.cell(row=row_idx, column=1)
        cell.value = row_idx - 2
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = border
        
        # Account No.
        cell = ws.cell(row=row_idx, column=2)
        cell.value = account.account_number
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Bank Name
        cell = ws.cell(row=row_idx, column=3)
        cell.value = account.bank_name
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # IFSC Code
        cell = ws.cell(row=row_idx, column=4)
        cell.value = account.ifsc_code if account.ifsc_code else ""
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Address
        cell = ws.cell(row=row_idx, column=5)
        cell.value = account.address if account.address else ""
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Mobile
        cell = ws.cell(row=row_idx, column=6)
        cell.value = account.mobile if account.mobile else ""
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Email
        cell = ws.cell(row=row_idx, column=7)
        cell.value = account.email if account.email else ""
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Transaction Count
        cell = ws.cell(row=row_idx, column=8)
        cell.value = account.transaction_count
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = border
        cell.number_format = '#,##0'
        
        # ACK Numbers
        cell = ws.cell(row=row_idx, column=9)
        cell.value = account.ack_numbers
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = border
        
        # Total Amount
        cell = ws.cell(row=row_idx, column=10)
        cell.value = account.total_amount
        cell.font = normal_font
        cell.alignment = right_align
        cell.border = border
        cell.number_format = '₹#,##0.00'
        
        # Disputed Amount
        cell = ws.cell(row=row_idx, column=11)
        cell.value = account.disputed_amount
        cell.font = normal_font
        cell.alignment = right_align
        cell.border = border
        cell.number_format = '₹#,##0.00'
        
        ws.row_dimensions[row_idx].height = 20
    
    # Adjust column widths
    column_widths = {
        'A': 8,   # SR NO
        'B': 20,  # Account No.
        'C': 30,  # Bank Name
        'D': 15,  # IFSC Code
        'E': 40,  # Address
        'F': 15,  # Mobile
        'G': 25,  # Email
        'H': 15,  # Transaction Count
        'I': 30,  # ACK Numbers
        'J': 18,  # Total Amount
        'K': 18   # Disputed Amount
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
