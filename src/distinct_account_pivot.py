import streamlit as st
import pandas as pd
import io
from datetime import datetime
import os
import json

def load_distinct_pivot_mappings():
    """Load column mappings from persistent JSON file"""
    mapping_file = "cyber multiple accoun with DA and ACK/.kiro/distinct_pivot_mappings.json"
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(mapping_file), exist_ok=True)
    
    if os.path.exists(mapping_file):
        try:
            with open(mapping_file, 'r') as f:
                return json.load(f)
        except:
            pass
    
    # Default mappings
    return {
        'bank_col': '-- Select --',
        'account_col': '-- Select --',
        'disputed_col': '-- Select --',
        'pivot_col': '-- Select --'
    }


def save_distinct_pivot_mappings(mappings):
    """Save column mappings to persistent JSON file"""
    mapping_file = "cyber multiple accoun with DA and ACK/.kiro/distinct_pivot_mappings.json"
    
    try:
        os.makedirs(os.path.dirname(mapping_file), exist_ok=True)
        with open(mapping_file, 'w') as f:
            json.dump(mappings, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Could not save mappings: {e}")
        return False


def render_distinct_account_pivot_page():
    """Render the Distinct Account Pivot page - creates pivot table from layerwise file."""
    
    st.title("📊 Distinct Account Pivot Report")
    st.markdown("Upload a layerwise file to generate a pivot table with bank-wise account counts and disputed amounts.")
    
    st.info("📌 This tool calculates total account count, distinct account count, and sum of disputed amount grouped by Bank Name and pivoted by a selected column (e.g., Layer).")
    
    # Initialize session state for column mappings
    if 'distinct_pivot_col_mappings' not in st.session_state:
        st.session_state.distinct_pivot_col_mappings = load_distinct_pivot_mappings()
    
    # File upload
    st.markdown("### 📁 Upload Layerwise File")
    uploaded_file = st.file_uploader(
        "Choose Layerwise Excel or CSV file",
        type=['xlsx', 'xls', 'csv'],
        help="Excel or CSV file containing Bank Name, Account Number, and Disputed Amount",
        key="distinct_pivot_file"
    )
    
    if uploaded_file is not None:
        try:
            # Step 1: Load data
            if uploaded_file.name.lower().endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            st.success(f"✅ Loaded file: {len(df):,} rows")
            
            with st.expander("👁️ Preview Data (First 10 rows)"):
                st.dataframe(df.head(10))
            
            # Step 2: Column Selection
            st.markdown("---")
            st.markdown("### 🎯 Step 2: Select Columns")
            
            col_sel1, col_sel2 = st.columns(2)
            
            headers = ["-- Select --"] + list(df.columns)
            
            with col_sel1:
                bank_col = st.selectbox(
                    "Bank Name Column (Rows):",
                    options=headers,
                    index=headers.index(st.session_state.distinct_pivot_col_mappings['bank_col']) if st.session_state.distinct_pivot_col_mappings['bank_col'] in headers else 0,
                    key="dp_bank_col"
                )
                
                account_col = st.selectbox(
                    "Account Number Column (Values):",
                    options=headers,
                    index=headers.index(st.session_state.distinct_pivot_col_mappings['account_col']) if st.session_state.distinct_pivot_col_mappings['account_col'] in headers else 0,
                    key="dp_account_col"
                )
            
            with col_sel2:
                disputed_col = st.selectbox(
                    "Disputed Amount Column (Values):",
                    options=headers,
                    index=headers.index(st.session_state.distinct_pivot_col_mappings['disputed_col']) if st.session_state.distinct_pivot_col_mappings['disputed_col'] in headers else 0,
                    key="dp_disputed_col"
                )
                
                pivot_col = st.selectbox(
                    "Pivot Column (Optional):",
                    options=headers,
                    index=headers.index(st.session_state.distinct_pivot_col_mappings['pivot_col']) if st.session_state.distinct_pivot_col_mappings['pivot_col'] in headers else 0,
                    key="dp_pivot_col",
                    help="Optional: Select a column to create horizontal headers (e.g., Layer). Leave as '-- Select --' for a simple summary by bank."
                )
            
            # Validation
            if bank_col == "-- Select --" or account_col == "-- Select --" or disputed_col == "-- Select --":
                st.warning("⚠️ Please select all required columns (Bank, Account, and Disputed Amount) to generate the pivot report.")
                return
            
            # Step 3: Generate Pivot
            st.markdown("---")
            if st.button("🚀 Generate Distinct Account Pivot", type="primary", use_container_width=True):
                # Save mappings
                mappings_to_save = {
                    'bank_col': bank_col,
                    'account_col': account_col,
                    'disputed_col': disputed_col,
                    'pivot_col': pivot_col
                }
                st.session_state.distinct_pivot_col_mappings = mappings_to_save
                save_distinct_pivot_mappings(mappings_to_save)
                
                with st.spinner("⏳ Generating pivot table with maximum precision..."):
                    # Clean data
                    # Use unique internal column names to avoid issues if same column is selected for multiple roles
                    df_pivot = pd.DataFrame()
                    df_pivot['bank'] = df[bank_col].copy()
                    df_pivot['account'] = df[account_col].copy()
                    df_pivot['disputed'] = df[disputed_col].copy()
                    
                    if pivot_col != "-- Select --":
                        df_pivot['pivot'] = df[pivot_col].copy()
                    
                    # 1. Robust numeric conversion for amount (Handle currency symbols, commas, and invalid values)
                    def clean_numeric(val):
                        if pd.isna(val) or val == '' or val is None:
                            return 0.0
                        try:
                            # Convert to string and remove common non-numeric characters
                            val_str = str(val).replace('₹', '').replace(',', '').replace(' ', '').strip()
                            # Handle negative amounts if they exist in parentheses like (100.00)
                            if val_str.startswith('(') and val_str.endswith(')'):
                                val_str = '-' + val_str[1:-1]
                            return float(val_str)
                        except:
                            return 0.0
                    
                    df_pivot['disputed'] = df_pivot['disputed'].apply(clean_numeric)
                    
                    # 2. String Normalization (Crucial for correct GroupBy/Pivot)
                    # We strip whitespace and upper case the bank name to group "SBI" and "sbi" together
                    df_pivot['bank'] = df_pivot['bank'].fillna('UNKNOWN').astype(str).str.strip().str.upper()
                    
                    if pivot_col != "-- Select --":
                        df_pivot['pivot'] = df_pivot['pivot'].fillna('N/A').astype(str).str.strip().str.upper()
                    
                    # 3. Account Number Cleaning (Ensuring distinct count is accurate)
                    # Convert to string, strip spaces. We want to treat '00123' and '123' as DIFFERENT if they are strings,
                    # but we must ensure we don't count empty strings as a unique account.
                    df_pivot['account'] = df_pivot['account'].fillna('').astype(str).str.strip()
                    
                    # FILTER: If account is blank, we shouldn't count it for "Distinct Accounts"
                    # We replace blank with NaN so nunique() ignores it by default
                    df_pivot.loc[df_pivot['account'] == '', 'account'] = pd.NA
                    
                    # Create Pivot/Grouped Table
                    if pivot_col != "-- Select --":
                        # Create Pivot Table with horizontal dimension
                        pivot_result = pd.pivot_table(
                            df_pivot,
                            index='bank',
                            columns='pivot',
                            values=['account', 'disputed'],
                            aggfunc={
                                'account': ['count', pd.Series.nunique], # count handles the total, nunique handles distinct
                                'disputed': 'sum'
                            },
                            fill_value=0
                        )
                        
                        # Flatten MultiIndex columns
                        columns_flat = []
                        for col in pivot_result.columns:
                            val_type, func, pivot_val = col
                            if func == 'count':
                                label = f"{pivot_val} - Count of Accounts"
                            elif func == 'nunique':
                                label = f"{pivot_val} - Distinct Accounts"
                            else: # sum
                                label = f"{pivot_val} - Sum of Disputed Amount"
                            columns_flat.append(label)
                        pivot_result.columns = columns_flat
                    else:
                        # Create simple grouping by bank
                        pivot_result = df_pivot.groupby('bank').agg({
                            'account': ['count', pd.Series.nunique],
                            'disputed': 'sum'
                        })
                        pivot_result.columns = ['Count of Account Number', 'Distinct Count of Account Number', 'Sum of Disputed Amount']
                    
                    # Round amounts to 2 decimal places for financial accuracy
                    amount_cols = [c for c in pivot_result.columns if 'Sum' in c or 'Amount' in c]
                    for col in amount_cols:
                        pivot_result[col] = pivot_result[col].round(2)
                    
                    # Add Total Row
                    if not pivot_result.empty:
                        numeric_pivot = pivot_result.apply(pd.to_numeric, errors='coerce').fillna(0)
                        total_row = numeric_pivot.sum().to_frame().T
                        total_row.index = ['TOTAL']
                        pivot_result = pd.concat([pivot_result, total_row])
                    
                    st.session_state.dp_pivot_result = pivot_result
                    st.success("✅ Pivot report generated successfully!")
            
            # Display and Download
            if 'dp_pivot_result' in st.session_state and st.session_state.dp_pivot_result is not None:
                pivot_result = st.session_state.dp_pivot_result
                
                st.markdown("### 📊 Pivot Report Results")
                st.dataframe(pivot_result, use_container_width=True)
                
                st.markdown("---")
                st.markdown("### 📥 Download Report")
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # Excel download
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    pivot_result.to_excel(writer, sheet_name='Distinct Account Pivot')
                    workbook = writer.book
                    worksheet = writer.sheets['Distinct Account Pivot']
                    
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                    
                    # Define Styles
                    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=11)
                    total_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
                    total_font = Font(bold=True)
                    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    center_align = Alignment(horizontal='center', vertical='center')
                    right_align = Alignment(horizontal='right', vertical='center')
                    left_align = Alignment(horizontal='left', vertical='center')

                    # Style Header
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                        cell.border = border

                    # Style Data Rows
                    num_rows = worksheet.max_row
                    num_cols = worksheet.max_column
                    
                    for row_idx in range(2, num_rows + 1):
                        is_total = (worksheet.cell(row=row_idx, column=1).value == "TOTAL")
                        for col_idx in range(1, num_cols + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = border
                            
                            if is_total:
                                cell.fill = total_fill
                                cell.font = total_font
                            
                            # Alignment
                            if col_idx == 1:
                                cell.alignment = left_align
                            else:
                                cell.alignment = right_align
                                if not is_total:
                                    # Alternating row color
                                    if row_idx % 2 == 1:
                                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                    # Auto-adjust column widths
                    worksheet.column_dimensions['A'].width = 35 # Bank Name
                    for idx, col in enumerate(pivot_result.columns, 2):
                        col_letter = worksheet.cell(row=1, column=idx).column_letter
                        worksheet.column_dimensions[col_letter].width = 25
                
                st.download_button(
                    label="⬇️ Download Pivot Report (Excel)",
                    data=excel_buffer.getvalue(),
                    file_name=f"distinct_account_pivot_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dp_download_excel"
                )
                
                # CSV download
                csv_data = pivot_result.to_csv().encode('utf-8')
                st.download_button(
                    label="⬇️ Download Pivot Report (CSV)",
                    data=csv_data,
                    file_name=f"distinct_account_pivot_{timestamp}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="dp_download_csv"
                )
        
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
    else:
        st.info("👆 Please upload a layerwise file to get started")
