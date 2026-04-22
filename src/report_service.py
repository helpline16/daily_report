"""
Report Service for generating Account, Hold Amount, and Unattended Complaint Reports.

This module generates professional Excel reports with proper formatting:
- Sky blue title rows
- Light yellow headers
- Table format with borders
- Grand total rows in bold with sky blue background
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import Dict, Tuple, List
import io


class ReportService:
    """Service for generating formatted Excel reports."""
    
    # Color definitions
    SKY_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    LIGHT_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Bright yellow
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Font definitions
    TITLE_FONT = Font(name='Calibri', size=14, bold=True, color="000000")
    HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="000000")
    NORMAL_FONT = Font(name='Calibri', size=12, color="000000")
    TOTAL_FONT = Font(name='Calibri', size=12, bold=True, color="000000")
    
    # Border definition
    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Alignment
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    
    def __init__(self):
        """Initialize the report service."""
        self.cleaning_log = []  # Track all cleaning operations
    
    def clean_layerwise_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean layerwise data - FIRST STEP before any processing.
        
        Args:
            df: Raw layerwise DataFrame
            
        Returns:
            Cleaned DataFrame with only valid entries
        """
        initial_count = len(df)
        self.cleaning_log.append(f"Initial layerwise records: {initial_count}")
        
        # Find ACK column
        ack_col = None
        for col in df.columns:
            if 'acknowledgement' in col.lower() or 'ack' in col.lower():
                ack_col = col
                break
        
        if ack_col is None:
            raise ValueError("Could not find Acknowledgement Number column in layerwise file")
        
        # Convert ACK to string and remove whitespace
        df[ack_col] = df[ack_col].astype(str).str.strip()
        
        # Remove entries where ACK doesn't start with "311"
        valid_mask = df[ack_col].str.startswith('311', na=False)
        removed_count = (~valid_mask).sum()
        
        df_clean = df[valid_mask].copy()
        
        self.cleaning_log.append(f"Removed {removed_count} entries where ACK doesn't start with '311'")
        self.cleaning_log.append(f"Valid layerwise records after cleaning: {len(df_clean)}")
        
        return df_clean
    
    def clean_hold_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean hold data - FIRST STEP before any processing.
        
        Args:
            df: Raw hold DataFrame
            
        Returns:
            Cleaned DataFrame with only valid entries
        """
        initial_count = len(df)
        self.cleaning_log.append(f"Initial hold records: {initial_count}")
        
        # Find ACK column
        ack_col = None
        for col in df.columns:
            if 'acknowledgement' in col.lower() or 'ack' in col.lower():
                ack_col = col
                break
        
        if ack_col is None:
            self.cleaning_log.append("Warning: No ACK column found in hold file, skipping ACK validation")
            return df
        
        # Convert ACK to string and remove whitespace
        df[ack_col] = df[ack_col].astype(str).str.strip()
        
        # Remove entries where ACK doesn't start with "311"
        valid_mask = df[ack_col].str.startswith('311', na=False)
        removed_count = (~valid_mask).sum()
        
        df_clean = df[valid_mask].copy()
        
        self.cleaning_log.append(f"Removed {removed_count} entries from hold file where ACK doesn't start with '311'")
        self.cleaning_log.append(f"Valid hold records after cleaning: {len(df_clean)}")
        
        return df_clean
    
    def clean_unattend_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean unattend data - FIRST STEP before any processing.
        
        Args:
            df: Raw unattend DataFrame
            
        Returns:
            Cleaned DataFrame with only valid entries
        """
        initial_count = len(df)
        self.cleaning_log.append(f"Initial unattend records: {initial_count}")
        
        # Find ACK column
        ack_col = None
        for col in df.columns:
            if 'acknowledgement' in col.lower() or 'ack' in col.lower():
                ack_col = col
                break
        
        if ack_col is None:
            self.cleaning_log.append("Warning: No ACK column found in unattend file, skipping ACK validation")
            return df
        
        # Convert ACK to string and remove whitespace
        df[ack_col] = df[ack_col].astype(str).str.strip()
        
        # Remove entries where ACK doesn't start with "311"
        valid_mask = df[ack_col].str.startswith('311', na=False)
        removed_count = (~valid_mask).sum()
        
        df_clean = df[valid_mask].copy()
        
        self.cleaning_log.append(f"Removed {removed_count} entries from unattend file where ACK doesn't start with '311'")
        self.cleaning_log.append(f"Valid unattend records after cleaning: {len(df_clean)}")
        
        return df_clean
    
    def generate_account_report(self, layerwise_df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate Account Report from layerwise data.
        
        Args:
            layerwise_df: CLEANED DataFrame with columns including 'Acknowledgement No.' and 'Bank/FIs'
            
        Returns:
            DataFrame with columns: SR NO, Bank, Count of Account No
        """
        # Validate required columns exist
        ack_col = None
        bank_col = None
        
        for col in layerwise_df.columns:
            if 'acknowledgement' in col.lower() or 'ack' in col.lower():
                ack_col = col
            if 'bank' in col.lower() or 'fi' in col.lower():
                bank_col = col
        
        if ack_col is None:
            raise ValueError("CRITICAL: Acknowledgement Number column not found in layerwise file")
        if bank_col is None:
            raise ValueError("CRITICAL: Bank/FIs column not found in layerwise file")
        
        self.cleaning_log.append(f"Using ACK column: {ack_col}")
        self.cleaning_log.append(f"Using Bank column: {bank_col}")
        
        # Filter banks - only include rows where Bank/FIs contains "Bank" or "bank" or "BANK"
        bank_df = layerwise_df[
            layerwise_df[bank_col].astype(str).str.contains('Bank|bank|BANK', na=False, regex=True)
        ].copy()
        
        self.cleaning_log.append(f"Filtered to {len(bank_df)} records containing 'Bank' in bank name")
        
        # Remove rows with empty or null ACK numbers
        bank_df = bank_df[bank_df[ack_col].notna() & (bank_df[ack_col].astype(str).str.strip() != '')].copy()
        
        # Group by Bank and count distinct Acknowledgement numbers
        account_counts = bank_df.groupby(bank_col)[ack_col].nunique().reset_index()
        account_counts.columns = ['Bank', 'Count of Account No']
        
        self.cleaning_log.append(f"Generated account report with {len(account_counts)} unique banks")
        
        # Sort by count descending (high to low)
        account_counts = account_counts.sort_values('Count of Account No', ascending=False).reset_index(drop=True)
        
        # Add SR NO column
        account_counts.insert(0, 'SR NO', range(1, len(account_counts) + 1))
        
        # Validate no missing data
        if account_counts['Count of Account No'].isna().any():
            raise ValueError("CRITICAL: Missing count values detected in account report")
        
        # Convert count to integer (no decimals allowed)
        account_counts['Count of Account No'] = account_counts['Count of Account No'].astype(int)
        
        return account_counts
    
    def generate_hold_report(self, hold_df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate Hold Amount Report.
        
        Args:
            hold_df: DataFrame with columns including 'Bank' and 'Amount'
            
        Returns:
            DataFrame with columns: SR NO, BANK, PUT ON HOLD AMOUNT
        """
        # Find the amount column - prioritize "Amount" column
        amount_col = None
        
        # First try exact match for "Amount"
        if 'Amount' in hold_df.columns:
            amount_col = 'Amount'
        else:
            # Try case-insensitive search
            for col in hold_df.columns:
                if col.lower().strip() == 'amount':
                    amount_col = col
                    break
        
        # If not found, search for columns with "amount" or "disputed"
        if amount_col is None:
            for col in hold_df.columns:
                if 'amount' in col.lower() or 'disputed' in col.lower():
                    amount_col = col
                    break
        
        if amount_col is None:
            raise ValueError("Could not find amount column in hold file. Expected 'Amount' column.")
        
        # Find bank column - prioritize "Bank" column
        bank_col = None
        
        # First try exact match for "Bank"
        if 'Bank' in hold_df.columns:
            bank_col = 'Bank'
        else:
            # Try case-insensitive search
            for col in hold_df.columns:
                if col.lower().strip() == 'bank':
                    bank_col = col
                    break
        
        # If not found, search for columns with "bank" or "fi"
        if bank_col is None:
            for col in hold_df.columns:
                if 'bank' in col.lower() or 'fi' in col.lower():
                    bank_col = col
                    break
        
        if bank_col is None:
            raise ValueError("Could not find bank column in hold file. Expected 'Bank' column.")
        
        # Convert amount to numeric
        hold_df[amount_col] = pd.to_numeric(hold_df[amount_col], errors='coerce')
        
        # Remove rows with NaN amounts or bank names
        hold_df_clean = hold_df[hold_df[amount_col].notna() & hold_df[bank_col].notna()].copy()
        
        # Group by Bank and sum amounts (this handles multiple entries for same bank)
        hold_amounts = hold_df_clean.groupby(bank_col, as_index=False)[amount_col].sum()
        hold_amounts.columns = ['BANK', 'PUT ON HOLD AMOUNT']
        
        # Sort by amount descending (high to low)
        hold_amounts = hold_amounts.sort_values('PUT ON HOLD AMOUNT', ascending=False).reset_index(drop=True)
        
        # Add SR NO column
        hold_amounts.insert(0, 'SR NO', range(1, len(hold_amounts) + 1))
        
        # Round amounts to 2 decimal places
        hold_amounts['PUT ON HOLD AMOUNT'] = hold_amounts['PUT ON HOLD AMOUNT'].round(2)
        
        return hold_amounts
    
    def generate_unattended_report(self, unattend_df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate Unattended Complaint Report.
        
        Args:
            unattend_df: DataFrame with bank/wallet/merchant and unattended count columns
            
        Returns:
            DataFrame with columns: SR NO, Bank/Wallet/Merchant, Complaint Un attended
        """
        # Find the bank/merchant column (usually first column or one with bank/wallet/merchant keyword)
        bank_col = None
        for col in unattend_df.columns:
            if any(keyword in col.lower() for keyword in ['bank', 'wallet', 'merchant', 'fi']):
                bank_col = col
                break
        
        if bank_col is None:
            # Use first column as bank column
            bank_col = unattend_df.columns[0]
        
        # Find the "Complaint Un attended" column - prioritize exact match
        count_col = None
        
        # First try exact match for "Complaint Un attended"
        for col in unattend_df.columns:
            if col.strip() == 'Complaint Un attended':
                count_col = col
                break
        
        # If not found, try case-insensitive match
        if count_col is None:
            for col in unattend_df.columns:
                if 'complaint' in col.lower() and 'unattend' in col.lower():
                    count_col = col
                    break
        
        # If still not found, try any column with "unattend"
        if count_col is None:
            for col in unattend_df.columns:
                if 'unattend' in col.lower():
                    count_col = col
                    break
        
        # Last resort - use column H (index 7) if it exists
        if count_col is None and len(unattend_df.columns) > 7:
            count_col = unattend_df.columns[7]  # Column H is index 7
        
        if count_col is None:
            raise ValueError("Could not find 'Complaint Un attended' column in unattend file")
        
        # Create report dataframe
        unattend_report = unattend_df[[bank_col, count_col]].copy()
        unattend_report.columns = ['Bank/Wallet/Merchant', 'Complaint Un attended']
        
        # Remove rows with dashes or empty bank names
        unattend_report = unattend_report[
            unattend_report['Bank/Wallet/Merchant'].notna() & 
            (unattend_report['Bank/Wallet/Merchant'].astype(str).str.strip() != '') &
            (~unattend_report['Bank/Wallet/Merchant'].astype(str).str.match(r'^-+$'))  # Remove rows with only dashes
        ]
        
        # Convert count to numeric
        unattend_report['Complaint Un attended'] = pd.to_numeric(
            unattend_report['Complaint Un attended'], errors='coerce'
        )
        
        # Remove rows where count is 0 or NaN
        unattend_report = unattend_report[
            (unattend_report['Complaint Un attended'] > 0) & 
            (unattend_report['Complaint Un attended'].notna())
        ]
        
        # Remove rows where bank name is "other", "others", "othar", or "total" (case insensitive)
        unattend_report = unattend_report[
            ~unattend_report['Bank/Wallet/Merchant'].str.lower().str.strip().isin(['other', 'others', 'othar', 'total', 'grand total', 'grandtotal'])
        ]
        
        # Sort by count descending (high to low)
        unattend_report = unattend_report.sort_values('Complaint Un attended', ascending=False).reset_index(drop=True)
        
        # Add SR NO column
        unattend_report.insert(0, 'SR NO', range(1, len(unattend_report) + 1))
        
        # Convert count to integer (round first if needed, then convert - no decimals allowed)
        unattend_report['Complaint Un attended'] = unattend_report['Complaint Un attended'].round(0).astype(int)
        
        return unattend_report
    
    def format_worksheet(self, ws, df: pd.DataFrame, title: str, date_str: str):
        """
        Apply professional formatting to a worksheet.
        
        Args:
            ws: openpyxl worksheet
            df: DataFrame to write
            title: Title for the report
            date_str: Date string to include in title
        """
        # Determine if this is a hold report (has decimal values)
        is_hold_report = 'HOLD' in title.upper()
        
        # Write title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"{date_str} {title}"
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.SKY_BLUE_FILL
        title_cell.alignment = self.CENTER_ALIGN
        title_cell.border = self.THIN_BORDER
        
        # Set row height for title (increased by 30%)
        ws.row_dimensions[1].height = 32.5  # 25 * 1.3
        
        # Write header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = self.HEADER_FONT
            cell.fill = self.LIGHT_YELLOW_FILL
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        # Set row height for header (increased by 30%)
        ws.row_dimensions[2].height = 26  # 20 * 1.3
        
        # Write data rows
        for row_idx, row_data in enumerate(df.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = self.NORMAL_FONT
                cell.fill = self.WHITE_FILL
                cell.border = self.THIN_BORDER
                
                # Center alignment for all cells
                cell.alignment = self.CENTER_ALIGN
                
                # Format numbers based on report type
                if isinstance(value, (int, float)):
                    if is_hold_report and isinstance(value, float):
                        cell.number_format = '#,##0.00'  # Decimal format for hold amounts
                    else:
                        cell.number_format = '#,##0'  # Integer format for counts
            
            # Set row height for data rows (increased by 30%)
            ws.row_dimensions[row_idx].height = 19.5  # 15 * 1.3
        
        # Add Grand Total row
        total_row_idx = len(df) + 3
        
        # Merge cells for "GRAND TOTAL" label
        ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=2)
        total_label_cell = ws.cell(row=total_row_idx, column=1)
        total_label_cell.value = "GRAND TOTAL"
        total_label_cell.font = self.TOTAL_FONT
        total_label_cell.fill = self.SKY_BLUE_FILL
        total_label_cell.alignment = self.CENTER_ALIGN
        total_label_cell.border = self.THIN_BORDER
        
        # Apply formatting to merged cells
        for col_idx in range(1, 3):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = self.TOTAL_FONT
            cell.fill = self.SKY_BLUE_FILL
            cell.border = self.THIN_BORDER
        
        # Calculate and write total for numeric column (last column)
        last_col_idx = len(df.columns)
        total_value = df.iloc[:, -1].sum()
        total_cell = ws.cell(row=total_row_idx, column=last_col_idx)
        total_cell.value = total_value
        total_cell.font = self.TOTAL_FONT
        total_cell.fill = self.SKY_BLUE_FILL
        total_cell.alignment = self.RIGHT_ALIGN
        total_cell.border = self.THIN_BORDER
        
        # Format total based on report type
        if is_hold_report:
            total_cell.number_format = '#,##0.00'  # Decimal format for hold total
        else:
            total_cell.number_format = '#,##0'  # Integer format for count totals
        
        # Set row height for total (increased by 30%)
        ws.row_dimensions[total_row_idx].height = 26  # 20 * 1.3
        
        # Auto-adjust column widths
        from openpyxl.utils import get_column_letter
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_length = len(str(col_name))
            for row_idx in range(3, len(df) + 4):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width (add some padding)
            column_letter = get_column_letter(col_idx)
            base_width = min(max_length + 3, 50)
            
            # Increase all columns by 20% except SR NO (column A)
            if col_idx != 1:  # Not SR NO column
                base_width = base_width * 1.2
            
            # Additional 40% increase for Bank column (middle column, column B)
            # Total: 20% + 40% = 60% wider for Bank column
            if col_idx == 2:  # Column B (Bank/Wallet/Merchant column)
                base_width = base_width * 1.4  # Additional 40% on top of the 20%
            
            ws.column_dimensions[column_letter].width = base_width
    
    def generate_complete_report(
        self,
        layerwise_df: pd.DataFrame,
        hold_df: pd.DataFrame,
        unattend_df: pd.DataFrame
    ) -> Tuple[bytes, List[str], Dict[str, float]]:
        """
        Generate complete report with all three sheets.
        
        CRITICAL: This function cleans data FIRST before any processing.
        
        Args:
            layerwise_df: Raw Layerwise/All Layer data
            hold_df: Raw Hold amount data
            unattend_df: Raw Unattended complaint data
            
        Returns:
            Tuple of (Excel file as bytes, cleaning log messages, grand totals dict)
        """
        # STEP 1: CLEAN ALL DATA FIRST - Remove entries where ACK doesn't start with "311"
        self.cleaning_log = []  # Reset log
        self.cleaning_log.append("=" * 60)
        self.cleaning_log.append("STEP 1: DATA CLEANING - Removing invalid ACK numbers")
        self.cleaning_log.append("=" * 60)
        
        layerwise_clean = self.clean_layerwise_data(layerwise_df)
        hold_clean = self.clean_hold_data(hold_df)
        unattend_clean = self.clean_unattend_data(unattend_df)
        
        self.cleaning_log.append("=" * 60)
        self.cleaning_log.append("STEP 2: GENERATING REPORTS FROM CLEANED DATA")
        self.cleaning_log.append("=" * 60)
        
        # STEP 2: Generate individual reports from CLEANED data
        account_report = self.generate_account_report(layerwise_clean)
        self.cleaning_log.append(f"✓ Account report generated: {len(account_report)} banks")
        
        hold_report = self.generate_hold_report(hold_clean)
        self.cleaning_log.append(f"✓ Hold report generated: {len(hold_report)} banks")
        
        unattended_report = self.generate_unattended_report(unattend_clean)
        self.cleaning_log.append(f"✓ Unattended report generated: {len(unattended_report)} entries")
        
        # Calculate grand totals
        grand_totals = {
            'hold_amount': float(hold_report['PUT ON HOLD AMOUNT'].sum()),
            'account_count': int(account_report['Count of Account No'].sum()),
            'unattended_count': int(unattended_report['Complaint Un attended'].sum())
        }
        
        # STEP 3: Validate reports have data
        if len(account_report) == 0:
            raise ValueError("CRITICAL: Account report is empty after processing")
        if len(hold_report) == 0:
            raise ValueError("CRITICAL: Hold report is empty after processing")
        if len(unattended_report) == 0:
            raise ValueError("CRITICAL: Unattended report is empty after processing")
        
        self.cleaning_log.append("=" * 60)
        self.cleaning_log.append("STEP 3: CREATING EXCEL FILE WITH FORMATTING")
        self.cleaning_log.append("=" * 60)
        
        # STEP 4: Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Get YESTERDAY's date
        from datetime import timedelta
        yesterday = datetime.now() - timedelta(days=1)
        date_str = yesterday.strftime("%d-%m-%Y")
        
        # Create and format PUT ON HOLD sheet
        ws_hold = wb.create_sheet("PUT ON HOLD")
        self.format_worksheet(ws_hold, hold_report, "PUT ON HOLD AMOUNT", date_str)
        self.cleaning_log.append("✓ PUT ON HOLD sheet formatted")
        
        # Create and format ACCOUNT sheet
        ws_account = wb.create_sheet("ACCOUNT")
        self.format_worksheet(ws_account, account_report, "Bank account details (involved in fraud)", date_str)
        self.cleaning_log.append("✓ ACCOUNT sheet formatted")
        
        # Create and format Complaint Un Attended sheet
        ws_unattend = wb.create_sheet(" Complaint Un Attended")
        self.format_worksheet(ws_unattend, unattended_report, "Unattended  Complaint", date_str)
        self.cleaning_log.append("✓ Complaint Un Attended sheet formatted")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        self.cleaning_log.append("=" * 60)
        self.cleaning_log.append("REPORT GENERATION COMPLETED SUCCESSFULLY")
        self.cleaning_log.append("=" * 60)
        
        return output.getvalue(), self.cleaning_log, grand_totals
